from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
GENERAL_DATA_DIR = ROOT / "SMT_2020 - Final" / "General Data"
OUTPUT_PATH = ROOT / "db" / "postgresql" / "010_create_smt2020_fab_staging_schemas.sql"

DATASETS = [
    ("dataset 1", "fab10", "SMT_2020_Model_Data_-_HVLM.xlsx"),
    ("dataset 2", "fab11", "SMT_2020_Model_Data_-_LVHM.xlsx"),
    ("dataset 3", "fab12", "SMT_2020_Model_Data_-_HVLM_E.xlsx"),
    ("dataset 4", "fab13", "SMT_2020_Model_Data_-_LVHM_E.xlsx"),
]

HEADER_ROWS = {
    "Setup_Matrix_Implant_Gas": 7,
}


def sql_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def snake(value: object, fallback: str) -> str:
    raw = str(value).strip() if value is not None else fallback
    raw = raw.replace("%", " percent ")
    raw = re.sub(r"[^0-9A-Za-z]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_").lower()
    if not raw:
        raw = fallback
    if raw[0].isdigit():
        raw = f"col_{raw}"
    return raw


def dedupe(names: Iterable[str]) -> list[str]:
    counts: Counter[str] = Counter()
    result: list[str] = []
    for name in names:
        counts[name] += 1
        result.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return result


def infer_pg_type(values: list[object]) -> str:
    non_empty = [value for value in values if value is not None and value != ""]
    if not non_empty:
        return "TEXT"
    if all(isinstance(value, datetime) for value in non_empty):
        return "TIMESTAMP"
    if all(isinstance(value, date) and not isinstance(value, datetime) for value in non_empty):
        return "DATE"
    if all(isinstance(value, bool) for value in non_empty):
        return "BOOLEAN"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in non_empty):
        return "INTEGER"
    numeric_types = (int, float, Decimal)
    if all(isinstance(value, numeric_types) and not isinstance(value, bool) for value in non_empty):
        return "NUMERIC"
    return "TEXT"


def sample_column_values(ws, header_row: int, column_index: int, limit: int = 500) -> list[object]:
    values: list[object] = []
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=min(ws.max_row, header_row + limit),
        min_col=column_index,
        max_col=column_index,
        values_only=True,
    ):
        values.append(row[0])
    return values


def table_ddl(schema_name: str, sheet_name: str, ws) -> str:
    header_row = HEADER_ROWS.get(sheet_name, 1)
    raw_headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    column_names = dedupe(
        snake(header, f"col_{idx:03d}") for idx, header in enumerate(raw_headers, start=1)
    )

    columns = [
        "    source_row_id BIGSERIAL PRIMARY KEY",
    ]
    for idx, column_name in enumerate(column_names, start=1):
        pg_type = infer_pg_type(sample_column_values(ws, header_row, idx))
        columns.append(f"    {ident(column_name)} {pg_type}")

    table_name = snake(sheet_name, "sheet")
    ddl = [
        f"CREATE TABLE IF NOT EXISTS {ident(schema_name)}.{ident(table_name)} (",
        ",\n".join(columns),
        ");",
        (
            f"COMMENT ON TABLE {ident(schema_name)}.{ident(table_name)} IS "
            f"{sql_quote(f'SMT2020 General Data sheet: {sheet_name}; header row: {header_row}')};"
        ),
    ]
    return "\n".join(ddl)


def main() -> None:
    lines = [
        "-- SMT2020 General Data staging schemas for PostgreSQL.",
        "-- Dataset mapping: dataset 1=fab10, dataset 2=fab11, dataset 3=fab12, dataset 4=fab13.",
        "-- Each Excel worksheet is represented as one table in its mapped fab schema.",
        "",
    ]

    for dataset_dir, schema_name, workbook_name in DATASETS:
        workbook_path = GENERAL_DATA_DIR / dataset_dir / workbook_name
        lines.append(f"-- {dataset_dir} -> {schema_name}: {workbook_path.relative_to(ROOT)}")
        lines.append(f"CREATE SCHEMA IF NOT EXISTS {ident(schema_name)};")
        lines.append("")

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                lines.append(table_ddl(schema_name, ws.title, ws))
                lines.append("")
        finally:
            wb.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
