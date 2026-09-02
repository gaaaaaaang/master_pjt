from __future__ import annotations

import argparse
import csv
import subprocess
import tempfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook as openpyxl_load_workbook

from generate_smt2020_postgres_ddl import (
    DATASETS,
    GENERAL_DATA_DIR,
    HEADER_ROWS,
    OUTPUT_PATH,
    PROJECT_ROOT,
    dedupe,
    ident,
    snake,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load SMT2020 General Data workbooks into fab10-fab13 PostgreSQL schemas."
    )
    parser.add_argument("--compose-file", default=str(PROJECT_ROOT / "docker-compose.yml"))
    parser.add_argument("--service", default="postgres")
    parser.add_argument("--database", default="fab")
    parser.add_argument("--user", default="fab_user")
    parser.add_argument("--skip-ddl", action="store_true")
    parser.add_argument("--no-truncate", action="store_true")
    parser.add_argument(
        "--schema",
        action="append",
        choices=["fab10", "fab11", "fab12", "fab13"],
        help="Load only one schema. Can be passed multiple times.",
    )
    parser.add_argument(
        "--max-rows-per-table",
        type=int,
        help="Optional smoke-test limit for each worksheet table.",
    )
    return parser.parse_args()


def psql_base(args: argparse.Namespace) -> list[str]:
    return [
        "docker",
        "compose",
        "-f",
        args.compose_file,
        "exec",
        "-T",
        args.service,
        "psql",
        "-U",
        args.user,
        "-d",
        args.database,
        "-v",
        "ON_ERROR_STOP=1",
    ]


def run_sql(args: argparse.Namespace, sql: str) -> None:
    subprocess.run([*psql_base(args), "-c", sql], check=True)


def run_sql_file(args: argparse.Namespace, path: Path) -> None:
    subprocess.run([*psql_base(args), "-f", "-"], input=path.read_bytes(), check=True)


def sheet_columns(ws, sheet_name: str) -> tuple[int, list[str]]:
    header_row = HEADER_ROWS.get(sheet_name, 1)
    raw_headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    return header_row, dedupe(
        snake(header, f"col_{idx:03d}") for idx, header in enumerate(raw_headers, start=1)
    )


def is_empty(row: Iterable[object]) -> bool:
    return all(value is None or value == "" for value in row)


def export_sheet_csv(ws, header_row: int, csv_path: Path, max_rows: int | None) -> int:
    written = 0
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row)
            if is_empty(values):
                continue
            writer.writerow(values)
            written += 1
            if max_rows is not None and written >= max_rows:
                break
    return written


def copy_csv(args: argparse.Namespace, schema_name: str, table_name: str, columns: list[str], csv_path: Path) -> None:
    column_sql = ", ".join(ident(column) for column in columns)
    copy_sql = (
        f"\\copy {ident(schema_name)}.{ident(table_name)} ({column_sql}) "
        "FROM STDIN WITH (FORMAT csv, NULL '')"
    )
    subprocess.run([*psql_base(args), "-c", copy_sql], input=csv_path.read_bytes(), check=True)


def load_dataset_workbook(
    args: argparse.Namespace, dataset_dir: str, schema_name: str, workbook_name: str
) -> None:
    workbook_path = GENERAL_DATA_DIR / dataset_dir / workbook_name
    print(f"\n[{dataset_dir} -> {schema_name}] {workbook_path.relative_to(PROJECT_ROOT)}")
    wb = openpyxl_load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        with tempfile.TemporaryDirectory(prefix="smt2020_csv_") as tmpdir:
            tmp = Path(tmpdir)
            for ws in wb.worksheets:
                table_name = snake(ws.title, "sheet")
                header_row, columns = sheet_columns(ws, ws.title)
                if not args.no_truncate:
                    run_sql(
                        args,
                        f"TRUNCATE TABLE {ident(schema_name)}.{ident(table_name)} RESTART IDENTITY;",
                    )
                csv_path = tmp / f"{schema_name}__{table_name}.csv"
                row_count = export_sheet_csv(ws, header_row, csv_path, args.max_rows_per_table)
                if row_count:
                    copy_csv(args, schema_name, table_name, columns, csv_path)
                print(f"  {schema_name}.{table_name}: {row_count} rows")
    finally:
        wb.close()


def main() -> None:
    args = parse_args()
    schemas = set(args.schema or ["fab10", "fab11", "fab12", "fab13"])

    if not args.skip_ddl:
        print(f"Applying DDL: {OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
        run_sql_file(args, OUTPUT_PATH)

    for dataset in DATASETS:
        if dataset[1] in schemas:
            load_dataset_workbook(args, *dataset)


if __name__ == "__main__":
    main()
